from __future__ import annotations

import csv
import io
from datetime import datetime
from typing import Any, Dict, Iterable, List, Tuple

from django.http import HttpResponse
from django.utils import timezone

try:
    from openpyxl import Workbook, load_workbook
except Exception:  # pragma: no cover - 测试导入时 openpyxl 可选
    Workbook = None  # type: ignore
    load_workbook = None  # type: ignore


TEMPLATE_COLUMNS: List[str] = [
    "user_id",
    "name",
    "gender",
    "wechat_id",
    "voice_part",
    "department",
    "class_name",
    "phone_number",
    "email",
    "dorm",
    "birthday",  # YYYY-MM-DD 格式
    "hometown",
    "ethnicity",
    "political_status",
    "political_affiliation",
    "is_specialty",
    "is_centralized",
    "position",
    "join_month",  # YYYY-MM 格式
    "graduate_month",  # YYYY-MM 格式
    "status",  # 可选值：ACTIVE / ALUMNI / INACTIVE
    "tier",
    "portfolio",
    "is_admin",
    "password",
]


def _to_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    if value is None:
        return False
    s = str(value).strip().lower()
    return s in {"1", "true", "yes", "y", "on", "是"}


def _normalize_row(raw: Dict[str, Any]) -> Dict[str, Any]:
    if "class" in raw and not raw.get("class_name"):
        raw["class_name"] = raw.get("class")

    for key in ("is_specialty", "is_centralized", "is_admin"):
        if key in raw:
            raw[key] = _to_bool(raw.get(key))

    cleaned: Dict[str, Any] = {}
    for k, v in raw.items():
        if isinstance(v, str):
            cleaned[k] = v.strip()
        else:
            cleaned[k] = v

    if cleaned.get("password") == "":
        cleaned.pop("password", None)

    bday = cleaned.get("birthday")
    if isinstance(bday, str) and bday:
        for sep in ("-", "/", "."):
            parts = bday.split(sep)
            if len(parts) == 3 and all(parts):
                try:
                    dt = datetime(int(parts[0]), int(parts[1]), int(parts[2]))
                    cleaned["birthday"] = dt.strftime("%Y-%m-%d")
                    break
                except Exception:
                    pass

    vp = cleaned.get("voice_part")
    if not vp:
        cleaned["voice_part"] = "Other"
    wx = cleaned.get("wechat_id")
    if not wx:
        cleaned["wechat_id"] = "请及时填写正确微信号"
    jm = cleaned.get("join_month")
    if not jm:
        cleaned["join_month"] = timezone.localdate().strftime("%Y-%m")
    gm = cleaned.get("graduate_month")
    if gm is None:
        cleaned["graduate_month"] = ""
    # 梯队为空时默认二队
    tr = cleaned.get("tier")
    if not tr:
        cleaned["tier"] = "二队"
    st = cleaned.get("status")
    if not st:
        cleaned["status"] = "ACTIVE"

    return cleaned


def parse_csv(file_bytes: bytes, annotate_empty: bool = False) -> List[Dict[str, Any]]:
    text = file_bytes.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    rows: List[Dict[str, Any]] = []
    for row in reader:
        filtered = {k: row.get(k) for k in TEMPLATE_COLUMNS if k in row}
        normalized = _normalize_row(filtered)
        if annotate_empty:
            empty_fields = []
            for k in TEMPLATE_COLUMNS:
                if k in filtered:
                    v = filtered.get(k)
                    if v is None or (isinstance(v, str) and v.strip() == ""):
                        empty_fields.append(k)
            if empty_fields:
                normalized["__empty_fields"] = empty_fields
        rows.append(normalized)
    return rows


def parse_xlsx(file_bytes: bytes, annotate_empty: bool = False) -> List[Dict[str, Any]]:
    if load_workbook is None:
        raise RuntimeError("openpyxl not available")
    stream = io.BytesIO(file_bytes)
    wb = load_workbook(stream, read_only=True, data_only=True)
    ws = wb.active
    headers: List[str] = []
    rows: List[Dict[str, Any]] = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            headers = [str(h).strip() if h is not None else "" for h in row]
            continue
        data: Dict[str, Any] = {}
        empty_fields: List[str] = []
        for idx, cell in enumerate(row):
            if idx < len(headers) and headers[idx] in TEMPLATE_COLUMNS:
                header = headers[idx]
                value = cell if cell is not None else ""
                data[header] = value
                if value is None or (isinstance(value, str) and value.strip() == ""):
                    empty_fields.append(header)
        normalized = _normalize_row(data)
        if annotate_empty and empty_fields:
            normalized["__empty_fields"] = empty_fields
        rows.append(normalized)
    return rows


def parse_bulk_file(
    filename: str, file_bytes: bytes, annotate_empty: bool = False
) -> Tuple[List[Dict[str, Any]], str]:
    name = (filename or "").lower()
    if name.endswith(".csv"):
        return parse_csv(file_bytes, annotate_empty=annotate_empty), "csv"
    if name.endswith(".xlsx"):
        return parse_xlsx(file_bytes, annotate_empty=annotate_empty), "xlsx"
    raise ValueError("仅支持 CSV 或 .xlsx 格式，请将 .xls 转换为 .xlsx 后上传。")


def build_export_workbook(members: Iterable[Any]) -> Any:
    """使用 TEMPLATE_COLUMNS 作为表头生成当前成员的 .xlsx 工作簿。"""
    if Workbook is None:
        raise RuntimeError("openpyxl not available")
    wb = Workbook()
    ws = wb.active
    ws.title = "MembersExport"
    ws.append(TEMPLATE_COLUMNS)

    def to_plain_text(value: Any) -> str:
        """将任意值转换为安全的纯文本字符串。

        - 保证按文本序列化，避免 Excel 自动识别数字或日期
        - 转义可能触发 Excel 公式求值的起始字符
        """
        s = "" if value is None else str(value)
        stripped = s.lstrip("\t ")
        if stripped[:1] in {"=", "+", "-", "@"} or s[:1] == "\t":
            return "'" + s
        return s

    for m in members:
        user = getattr(m, "user", None)
        user_id = getattr(user, "user_id", "")
        role = getattr(user, "role", "")
        raw_row = [
            user_id,
            getattr(m, "name", ""),
            getattr(m, "gender", ""),
            getattr(m, "wechat_id", ""),
            getattr(m, "voice_part", ""),
            getattr(m, "department", ""),
            getattr(m, "class_name", ""),
            getattr(m, "phone_number", ""),
            getattr(m, "email", ""),
            getattr(m, "dorm", ""),
            getattr(m, "birthday", "") or "",  # 可能是 date 或 str，后续统一转换
            getattr(m, "hometown", ""),
            getattr(m, "ethnicity", ""),
            getattr(m, "political_status", ""),
            getattr(m, "political_affiliation", ""),
            "是" if getattr(m, "is_specialty", False) else "否",
            "是" if getattr(m, "is_centralized", False) else "否",
            getattr(m, "position", ""),
            getattr(m, "join_month", ""),
            getattr(m, "graduate_month", ""),
            getattr(m, "status", ""),
            getattr(m, "tier", ""),
            getattr(m, "portfolio", ""),
            "是" if role in ("Admin", "SuperAdmin") else "否",
            "",  # 不导出密码
        ]
        safe_row = [
            to_plain_text(v if i != 10 else (str(v) if v else ""))
            for i, v in enumerate(raw_row)
        ]
        ws.append(safe_row)
        row_idx = ws.max_row
        for cell in ws[row_idx]:
            cell.number_format = "@"
            cell.data_type = "s"
    return wb


def build_export_response(members: Iterable[Any]) -> HttpResponse:
    wb = build_export_workbook(members)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    resp = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = 'attachment; filename="members_export.xlsx"'
    return resp


def build_template_workbook() -> Any:
    if Workbook is None:
        raise RuntimeError("openpyxl not available")
    wb = Workbook()
    ws = wb.active
    ws.title = "MembersImport"
    ws.append(TEMPLATE_COLUMNS)
    ws.append(
        [
            "2025123456",
            "张三",
            "男",
            "weixin123",
            "T1",
            "000 建筑学院",
            "建15",
            "13812345678",
            "zhangsan@example.com",
            "紫荆#11-101",
            "2001-09-01",
            "北京",
            "汉族",
            "共青团员",
            "艺术团",
            "否",
            "否",
            "队员",
            "2025-09",
            "2029-07",
            "二队",
            "热爱合唱",
            "否",
            "ChangeMe123!",
        ]
    )
    return wb


def build_template_response() -> HttpResponse:
    wb = build_template_workbook()
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    resp = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = 'attachment; filename="members_import_template.xlsx"'
    return resp
